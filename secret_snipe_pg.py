"""
SecretSnipe Core Scanner Module - PostgreSQL/Redis Version

This module provides the core secret scanning functionality for the SecretSnipe platform.
It handles file processing, text extraction, signature matching, and PostgreSQL/Redis database operations.

Key Features:
- Multi-format file support (PDF, Excel, Word, images, archives)
- Advanced text extraction with OCR capabilities
- High-performance regex matching with Hyperscan support
- PostgreSQL for structured data storage
- Redis for caching and session management
- Parallel processing for large-scale scanning
- Comprehensive signature validation and compilation

Supported File Types:
- Plain text files (.txt, .py, .js, .java, etc.)
- PDF documents with OCR fallback
- Microsoft Office files (.xlsx, .xls, .docx, .doc)
- Images with OCR (.jpg, .png, .bmp, .tiff)
- Archive files (.zip) with recursive extraction

Database Features:
- PostgreSQL for structured data with full ACID compliance
- Redis for high-performance caching and session management
- Automatic schema management and migrations
- Connection pooling for concurrent access
- Comprehensive audit logging

Dependencies:
- psycopg2-binary for PostgreSQL connectivity
- redis for Redis operations
- PyMuPDF (fitz) for PDF processing
- openpyxl/xlrd for Excel files
- python-docx for Word documents
- EasyOCR for image text extraction
- Hyperscan for high-performance regex (optional)
- xxhash for fast file hashing
"""

import argparse
import json
import logging
import re
from pathlib import Path
import fitz  # PyMuPDF
import openpyxl
import xlrd
from docx import Document
from zipfile import ZipFile, BadZipFile
from itertools import groupby
from PIL import Image
import easyocr
try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
except ImportError:
    PYTESSERACT_AVAILABLE = False
    logging.warning("pytesseract not available, will use EasyOCR for image OCR")
try:
    import hyperscan
except ImportError:
    hyperscan = None
    logging.warning("Hyperscan not available, falling back to standard regex")
from tqdm import tqdm
from functools import partial
from datetime import datetime
import os
import sys
import string
import time
import psutil
import cProfile
import tempfile
import subprocess
import hashlib
import xxhash
import gc
from joblib import Parallel, delayed
from concurrent.futures import ThreadPoolExecutor, as_completed
from multiprocessing import cpu_count
import io
import numpy as np

# New imports for PostgreSQL and Redis
from database_manager import (
    db_manager, project_manager, scan_session_manager,
    findings_manager, file_cache_manager, init_database
)
from redis_manager import redis_manager, cache_manager, scan_cache
from config import config

# --- 1. CONFIGURATION & LOGGING ---
Image.MAX_IMAGE_PIXELS = None
logging.basicConfig(
    level=getattr(logging, config.log_level),
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(config.log_file, mode='a'),
        logging.StreamHandler()
    ]
)

# Global variables
reader = None  # EasyOCR reader (lazy loaded)
ocr_image_count = 0  # Counter for OCR processed images (for memory management)
signatures = []  # Compiled signatures

def load_signatures():
    """Load and compile signatures from JSON file"""
    global signatures
    try:
        with open("signatures.json", "r") as f:
            raw_signatures = json.load(f)

        # Validate and compile signatures
        valid_signatures = []
        for i, sig in enumerate(raw_signatures):
            try:
                # Validate required fields
                if not all(key in sig for key in ['name', 'regex', 'severity']):
                    logging.warning(f"Signature {i} missing required fields, skipping")
                    continue

                # Compile regex with error handling
                sig["regex"] = re.compile(sig["regex"])
                valid_signatures.append(sig)
                logging.debug(f"Compiled signature: {sig['name']}")
            except re.error as e:
                logging.warning(f"Invalid regex in signature {i}: {e}")
                continue

        signatures = valid_signatures
        logging.info(f"Loaded {len(signatures)} signatures")
        return True

    except FileNotFoundError:
        logging.error("signatures.json not found")
        return False
    except json.JSONDecodeError as e:
        logging.error(f"Error parsing signatures.json: {e}")
        return False

def get_ocr_reader():
    """Lazy load EasyOCR reader with memory management"""
    global reader, ocr_image_count
    
    # If OCR is disabled, return None immediately
    if not config.scanner.enable_ocr:
        return None
    
    # Reset OCR reader every N images to prevent memory accumulation
    max_ocr_images = int(os.getenv('OCR_RESET_AFTER_IMAGES', '50'))
    if reader is not None and ocr_image_count >= max_ocr_images:
        logging.info(f"Resetting OCR reader after {ocr_image_count} images to free memory")
        del reader
        reader = None
        ocr_image_count = 0
        gc.collect()
        
    if reader is None:
        try:
            # Ensure EasyOCR model directory exists and is writable
            easyocr_dir = os.path.expanduser("~/.EasyOCR")
            model_dir = os.path.join(easyocr_dir, "model")
            os.makedirs(model_dir, exist_ok=True)
            
            logging.info("Initializing EasyOCR reader...")
            # Initialize with proper model directory and error handling
            reader = easyocr.Reader(
                config.scanner.ocr_languages,
                model_storage_directory=model_dir,
                download_enabled=True
            )
            logging.info("EasyOCR reader initialized successfully")
        except Exception as e:
            logging.error(f"Failed to initialize EasyOCR: {e}")
            logging.warning("OCR functionality will be disabled for this scan")
            reader = False  # Use False instead of None to indicate permanent failure
    return reader if reader is not False else None

def extract_text_from_file(file_path):
    """Extract text from various file formats"""
    file_path = Path(file_path)
    file_extension = file_path.suffix.lower()

    try:
        if file_extension in ['.txt', '.py', '.js', '.ts', '.java', '.cpp', '.c', '.h',
                             '.php', '.rb', '.go', '.rs', '.swift', '.kt', '.scala',
                             '.clj', '.hs', '.ml', '.md', '.json', '.xml', '.yaml',
                             '.yml', '.toml', '.ini', '.cfg', '.conf', '.properties',
                             '.sh', '.bat', '.ps1', '.sql', '.html', '.css', '.scss']:
            # Plain text files
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()

        elif file_extension in ['.pdf']:
            # PDF files
            text = ""
            with fitz.open(file_path) as doc:
                for page in doc:
                    text += page.get_text()
            return text

        elif file_extension in ['.xlsx', '.xls']:
            # Excel files
            text = ""
            if file_extension == '.xlsx':
                workbook = openpyxl.load_workbook(file_path, read_only=True)
            else:
                workbook = xlrd.open_workbook(file_path)

            for sheet_name in workbook.sheetnames:
                if file_extension == '.xlsx':
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows(values_only=True):
                        text += ' '.join(str(cell) for cell in row if cell) + '\n'
                else:
                    sheet = workbook.sheet_by_name(sheet_name)
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        text += ' '.join(str(cell) for cell in row if cell) + '\n'
            return text

        elif file_extension in ['.docx']:
            # Word documents
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + '\n'
            return text

        elif file_extension in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif']:
            # Image files with OCR
            global ocr_image_count
            
            # Check which OCR engine to use (pytesseract is lower memory, easyocr is more accurate)
            use_easyocr = os.getenv('OCR_ENGINE', 'pytesseract').lower() == 'easyocr'
            
            # Skip OCR for very large images to prevent OOM
            # For pytesseract we can handle larger files since it uses external process
            if use_easyocr:
                max_ocr_size_mb = float(os.getenv('MAX_OCR_FILE_SIZE_MB', '0.5'))  # 500KB for EasyOCR
            else:
                max_ocr_size_mb = float(os.getenv('MAX_OCR_FILE_SIZE_MB', '5.0'))  # 5MB for pytesseract
            
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > max_ocr_size_mb:
                logging.info(f"Skipping OCR for large image {file_path} ({file_size_mb:.2f}MB > {max_ocr_size_mb}MB)")
                return ""
            
            # Try pytesseract first (low memory usage via external process)
            if PYTESSERACT_AVAILABLE and not use_easyocr:
                try:
                    # Use PIL to open image, pytesseract handles the rest externally
                    img = Image.open(file_path)
                    # Set timeout to prevent hanging on complex images
                    text = pytesseract.image_to_string(img, timeout=30)
                    img.close()
                    ocr_image_count += 1
                    return text.strip()
                except RuntimeError as e:
                    # Timeout or other runtime error
                    logging.warning(f"pytesseract timeout for {file_path}: {e}")
                    return ""
                except Exception as e:
                    logging.warning(f"pytesseract failed for {file_path}: {e}, trying EasyOCR fallback")
                    # Fall through to EasyOCR
            
            # Fallback to EasyOCR (more accurate but higher memory)
            reader = get_ocr_reader()
            if reader:
                try:
                    results = reader.readtext(str(file_path))
                    text = ' '.join([result[1] for result in results])
                    ocr_image_count += 1  # Increment counter for memory management
                    # Force cleanup
                    del results
                    gc.collect()
                    return text
                except Exception as e:
                    logging.error(f"OCR processing failed for {file_path}: {e}")
                    return ""
            else:
                logging.warning(f"OCR not available for {file_path}")
                return ""

        elif file_extension in ['.zip']:
            # ZIP archives
            text = ""
            try:
                with ZipFile(file_path, 'r') as zip_file:
                    for file_info in zip_file.filelist:
                        if not file_info.is_dir():
                            with zip_file.open(file_info) as f:
                                content = f.read().decode('utf-8', errors='ignore')
                                text += content + '\n'
            except BadZipFile:
                logging.warning(f"Invalid ZIP file: {file_path}")
            return text

        else:
            # Unknown file type
            logging.debug(f"Unsupported file type: {file_extension}")
            return ""

    except Exception as e:
        logging.warning(f"Error extracting text from {file_path}: {e}")
        return ""

def luhn_checksum(card_number: str) -> bool:
    """
    Validate a credit card number using the Luhn algorithm (mod 10 checksum).
    
    The Luhn algorithm:
    1. From the rightmost digit, double every second digit
    2. If doubling results in a number > 9, subtract 9
    3. Sum all the digits
    4. If the total modulo 10 equals 0, the number is valid
    
    Args:
        card_number: The card number string (may contain spaces/dashes)
        
    Returns:
        True if the card number passes Luhn validation, False otherwise
    """
    # Remove spaces, dashes, and any other non-digit characters
    digits = ''.join(c for c in card_number if c.isdigit())
    
    if not digits or len(digits) < 13 or len(digits) > 19:
        return False
    
    # Convert to list of integers, reversed
    digits_list = [int(d) for d in digits][::-1]
    
    # Apply Luhn algorithm
    checksum = 0
    for i, digit in enumerate(digits_list):
        if i % 2 == 1:  # Every second digit from right (index 1, 3, 5...)
            digit *= 2
            if digit > 9:
                digit -= 9
        checksum += digit
    
    return checksum % 10 == 0


def is_test_credit_card(card_number: str) -> bool:
    """
    Detect test/dummy credit card numbers that are clearly not real.
    
    Patterns detected:
    - Repeating groups (e.g., 5575 5575 5575 5575, 4000 4000 4000 4000)
    - Sequential numbers (e.g., 1234 5678 9012 3456)
    - All same digits (e.g., 4444 4444 4444 4444)
    - Known test card numbers from payment processors
    
    Args:
        card_number: The card number string
        
    Returns:
        True if the card appears to be test data, False otherwise
    """
    # Remove spaces, dashes, and any other non-digit characters
    digits = ''.join(c for c in card_number if c.isdigit())
    
    if len(digits) < 13:
        return False
    
    # Check for repeating 4-digit groups (e.g., 5575 5575 5575 5575)
    if len(digits) == 16:
        groups = [digits[i:i+4] for i in range(0, 16, 4)]
        if len(set(groups)) == 1:
            return True  # All 4 groups are identical
        # Check for 2-group repeat pattern (e.g., 1234 5678 1234 5678)
        if groups[0] == groups[2] and groups[1] == groups[3]:
            return True
    
    # Check for all same digits (e.g., 4444444444444444)
    if len(set(digits)) == 1:
        return True
    
    # Check for sequential patterns
    sequential_asc = ''.join(str(i % 10) for i in range(len(digits)))
    sequential_desc = ''.join(str((9 - i) % 10) for i in range(len(digits)))
    if digits == sequential_asc[:len(digits)] or digits == sequential_desc[:len(digits)]:
        return True
    
    # Known test card numbers from major payment processors
    known_test_cards = {
        # Stripe test cards
        '4242424242424242',  # Visa
        '4000056655665556',  # Visa (debit)
        '5555555555554444',  # Mastercard
        '2223003122003222',  # Mastercard (2-series)
        '5200828282828210',  # Mastercard (debit)
        '378282246310005',   # Amex
        '371449635398431',   # Amex
        '6011111111111117',  # Discover
        '6011000990139424',  # Discover
        # PayPal test cards
        '4111111111111111',  # Visa
        '4012888888881881',  # Visa
        # Braintree test cards
        '4009348888881881',
        '4012000033330026',
        '4012000077777777',
        # Authorize.net test cards
        '370000000000002',   # Amex
        '6011000000000012',  # Discover
        # Generic test patterns
        '0000000000000000',
        '1111111111111111',
        '1234567890123456',
    }
    
    if digits in known_test_cards:
        return True
    
    return False


def is_likely_not_credit_card(value: str, context: str) -> tuple:
    """
    Additional heuristics to detect false positive credit card numbers.
    
    Checks for patterns that look like credit cards but are actually:
    - Version numbers (1.2.3.4, v4.5.6.7)
    - Timestamps/dates (20241231123456)
    - UUIDs with numbers
    - Phone numbers
    - Hex color codes
    - IP addresses with ports
    - Build/revision numbers
    - Object IDs, sequence numbers
    
    Args:
        value: The matched credit card-like number
        context: Surrounding text context
        
    Returns:
        Tuple of (is_likely_false_positive, reason)
    """
    import re
    
    digits = ''.join(c for c in value if c.isdigit())
    context_lower = context.lower()
    
    # Check if the context suggests this is NOT a credit card
    non_card_context = [
        'version', 'v.', 'build', 'revision', 'rev.', 'release',
        'timestamp', 'datetime', 'date', 'time', 
        'id:', 'id=', 'id ', 'identifier', 'uid', 'uuid',
        'phone', 'tel', 'fax', 'mobile',
        'order', 'invoice', 'ref', 'reference', 'tracking',
        'serial', 'sn:', 'part', 'model', 'sku',
        'ip:', 'address:', 'port', 'sequence', 'seq',
        'index', 'offset', 'position', 'count',
        'checksum', 'hash', 'digest', 'crc',
    ]
    
    for marker in non_card_context:
        if marker in context_lower:
            # But not if card context is also present
            card_context = ['credit', 'card', 'payment', 'visa', 'mastercard', 
                           'amex', 'discover', 'pan', 'ccn', 'cardnum']
            has_card_context = any(cc in context_lower for cc in card_context)
            if not has_card_context:
                return True, f"Context suggests non-card data: '{marker}'"
    
    # Check for version number patterns (digits with dots nearby)
    if re.search(r'v?\d+\.\d+', context, re.IGNORECASE):
        return True, "Appears to be version number"
    
    # Check for timestamp-like patterns (year prefix)
    if digits[:4] in ['2024', '2025', '2026', '2023', '2022', '2021', '2020', '2019']:
        if len(digits) >= 14:  # Timestamp: YYYYMMDDHHMMSS
            return True, "Appears to be timestamp"
    
    # Check for year suffix (common in IDs)
    if digits[-4:] in ['2024', '2025', '2026', '2023', '2022', '2021', '2020']:
        if re.search(r'id|ref|order|invoice|tracking', context_lower):
            return True, "Appears to be dated reference number"
    
    # Check if all digit groups are the same length (often IDs, not cards)
    parts = re.findall(r'\d+', value)
    if len(parts) >= 3:
        lengths = [len(p) for p in parts]
        if len(set(lengths)) == 1 and lengths[0] <= 3:
            return True, "Appears to be formatted ID (uniform groups)"
    
    # Check for hex-like patterns nearby suggesting UUID/hash
    if re.search(r'[0-9a-f]{8}-?[0-9a-f]{4}', context, re.IGNORECASE):
        return True, "Appears to be part of UUID/hash"
    
    # Check for file path patterns that might contain version-like numbers
    if re.search(r'[\\/]v?\d+[\\/]', context) or re.search(r'_\d{13,16}_', context):
        return True, "Appears to be file path component or sequence"
    
    # If the match is mostly zeros (placeholder patterns)
    zero_count = digits.count('0')
    if zero_count >= len(digits) * 0.7:  # 70% or more zeros
        return True, "Too many zeros (likely placeholder)"
    
    return False, ""


def validate_ssn(ssn: str) -> tuple:
    """
    Validate a Social Security Number using known rules.
    
    SSN validation rules:
    - Cannot be all zeros in any group (000-XX-XXXX, XXX-00-XXXX, XXX-XX-0000)
    - Area number (first 3 digits) cannot be 000, 666, or 900-999
    - Cannot be known invalid/advertisement SSNs
    
    Args:
        ssn: The SSN string (with or without dashes)
        
    Returns:
        Tuple of (is_valid, reason) where reason explains invalidation
    """
    # Remove spaces and dashes
    digits = ''.join(c for c in ssn if c.isdigit())
    
    if len(digits) != 9:
        return False, "Invalid length"
    
    area = int(digits[:3])
    group = int(digits[3:5])
    serial = int(digits[5:])
    
    # Area number cannot be 000, 666, or 900-999
    if area == 0:
        return False, "Invalid area number (000)"
    if area == 666:
        return False, "Invalid area number (666)"
    if area >= 900:
        return False, "Invalid area number (900-999)"
    
    # Group number cannot be 00
    if group == 0:
        return False, "Invalid group number (00)"
    
    # Serial number cannot be 0000
    if serial == 0:
        return False, "Invalid serial number (0000)"
    
    # Known invalid SSNs (used in advertisements, etc.)
    known_invalid = {
        '078051120',  # Woolworth wallet SSN (most misused)
        '219099999',  # Used in advertisements
        '457555462',  # Used in commercials
        '123456789',  # Obviously fake
        '111111111',  # All same digits
        '222222222',
        '333333333',
        '444444444',
        '555555555',
        '666666666',
        '777777777',
        '888888888',
        '999999999',
    }
    
    if digits in known_invalid:
        return False, f"Known invalid/test SSN"
    
    # Check for repeating patterns
    if len(set(digits)) <= 2:
        return False, "Suspicious repeating pattern"
    
    return True, ""


def validate_aws_access_key(key: str) -> tuple:
    """
    Validate AWS Access Key ID format.
    
    AWS Access Key IDs:
    - Start with 'AKIA' (IAM user) or 'ASIA' (temporary/STS)
    - Followed by 16 alphanumeric characters
    - Total length: 20 characters
    
    Returns:
        Tuple of (is_valid, reason)
    """
    # Remove any whitespace
    key = key.strip()
    
    # Must be exactly 20 characters
    if len(key) != 20:
        return False, f"Invalid length: {len(key)} (expected 20)"
    
    # Must start with AKIA (long-term) or ASIA (temporary)
    if not key.startswith(('AKIA', 'ASIA', 'AIDA', 'AROA', 'AIPA', 'ANPA', 'ANVA', 'AGPA')):
        return False, "Invalid prefix (not a valid AWS key type)"
    
    # Rest must be alphanumeric (uppercase and digits)
    if not key[4:].isalnum():
        return False, "Contains invalid characters"
    
    # Check for obvious test patterns
    test_patterns = {
        'AKIAIOSFODNN7EXAMPLE',  # AWS documentation example
        'AKIAI44QH8DHBEXAMPLE',  # Another AWS example
        'AKIAXXXXXXXXXXXXXXXX',  # Placeholder pattern
        'AKIA0000000000000000',  # Zeros pattern
    }
    if key in test_patterns:
        return False, "Known test/example key"
    
    # Check for repeating patterns in the suffix
    suffix = key[4:]
    if len(set(suffix)) <= 2:
        return False, "Suspicious repeating pattern"
    
    return True, ""


def validate_jwt_token(token: str) -> tuple:
    """
    Validate JWT token structure and check for obvious test tokens.
    
    JWT structure: header.payload.signature (base64url encoded)
    
    Returns:
        Tuple of (is_valid, is_expired, reason)
    """
    import base64
    import json
    
    parts = token.split('.')
    if len(parts) != 3:
        return False, "Invalid JWT structure (not 3 parts)"
    
    try:
        # Decode header
        header_b64 = parts[0] + '=' * (4 - len(parts[0]) % 4)  # Add padding
        header = json.loads(base64.urlsafe_b64decode(header_b64))
        
        # Check for valid algorithm
        alg = header.get('alg', '')
        valid_algs = ['HS256', 'HS384', 'HS512', 'RS256', 'RS384', 'RS512', 
                      'ES256', 'ES384', 'ES512', 'PS256', 'PS384', 'PS512']
        if alg not in valid_algs and alg != 'none':
            return False, f"Unknown algorithm: {alg}"
        
        # Decode payload
        payload_b64 = parts[1] + '=' * (4 - len(parts[1]) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        
        # Check for test/example payloads
        sub = payload.get('sub', '')
        if sub in ['1234567890', 'test', 'example', 'user', 'admin']:
            return False, f"Likely test token (sub: {sub})"
        
        # Check if expired (if exp claim exists)
        exp = payload.get('exp')
        if exp:
            import time
            if exp < time.time():
                return False, "Token is expired"
        
        return True, ""
        
    except (json.JSONDecodeError, ValueError, TypeError) as e:
        return False, f"Invalid base64/JSON: {str(e)}"


def validate_private_key(key: str) -> tuple:
    """
    Validate that a private key block is not a test/example key.
    
    Returns:
        Tuple of (is_valid, reason)
    """
    # Check for obviously fake/test keys
    test_markers = [
        'EXAMPLE', 'example', 'TEST', 'test', 'FAKE', 'fake',
        'SAMPLE', 'sample', 'DUMMY', 'dummy', 'PLACEHOLDER'
    ]
    
    for marker in test_markers:
        if marker in key:
            return False, f"Contains test marker: {marker}"
    
    # Check key length (real keys have substantial content)
    key_content = key.replace('-', '').replace(' ', '').replace('\n', '')
    if len(key_content) < 100:
        return False, "Key content too short"
    
    return True, ""


def validate_database_url(url: str) -> tuple:
    """
    Validate database connection string and check for test credentials.
    
    Returns:
        Tuple of (is_valid, reason)
    """
    url_lower = url.lower()
    
    # Check for localhost/test hosts (less critical)
    test_hosts = ['localhost', '127.0.0.1', '0.0.0.0', 'example.com', 'test.local']
    for host in test_hosts:
        if host in url_lower:
            return False, f"Test/local host: {host}"
    
    # Check for obvious test credentials
    test_creds = ['password', 'admin:admin', 'root:root', 'test:test', 
                  'user:password', 'postgres:postgres', 'sa:sa']
    for cred in test_creds:
        if cred in url_lower:
            return False, f"Obvious test credential: {cred}"
    
    return True, ""


def validate_api_token_entropy(token: str) -> tuple:
    """
    Check if an API token has sufficient entropy (randomness).
    Low entropy tokens are likely test/placeholder values.
    
    Returns:
        Tuple of (is_valid, reason)
    """
    import math
    
    # Remove common prefixes
    for prefix in ['sk_', 'pk_', 'api_', 'key_', 'token_']:
        if token.lower().startswith(prefix):
            token = token[len(prefix):]
            break
    
    if len(token) < 10:
        return False, "Token too short"
    
    # Calculate character frequency
    freq = {}
    for char in token:
        freq[char] = freq.get(char, 0) + 1
    
    # Calculate Shannon entropy
    entropy = 0.0
    for count in freq.values():
        p = count / len(token)
        if p > 0:
            entropy -= p * math.log2(p)
    
    # Real tokens typically have entropy > 3.5 bits per character
    # Test tokens like 'xxxxxx' or 'abcdef' have low entropy
    if entropy < 2.5:
        return False, f"Low entropy ({entropy:.2f} bits) - likely placeholder"
    
    # Check for repeating patterns
    if len(set(token)) <= 3:
        return False, "Too few unique characters"
    
    return True, ""


def is_in_comment(context, position):
    """Check if a match is within a comment"""
    # Simple comment detection - can be enhanced
    lines = context.split('\n')
    if len(lines) < 3:
        return False

    # Check surrounding lines for comment markers
    for line in lines:
        line = line.strip()
        if line.startswith('//') or line.startswith('#') or line.startswith('/*') or '*/' in line:
            return True

    return False

def scan_text_with_signatures(text, file_path_str):
    """Scan text for secrets using compiled signatures"""
    findings = []

    for sig in signatures:
        try:
            for match in sig["regex"].finditer(text):
                value = match.group(0)

                # Get context around the match
                start = max(0, match.start() - 50)
                end = min(len(text), match.end() + 50)
                context = text[start:end]
                
                # For low-risk/context-based findings (like Confidentiality Markers),
                # provide more useful secret_value by including surrounding context
                if sig.get("low_risk", False) or sig["name"] in ["Confidentiality Marker", "Private IP Address"]:
                    # Get a snippet around the match for better context in the value
                    snippet_start = max(0, match.start() - 20)
                    snippet_end = min(len(text), match.end() + 20)
                    secret_display = text[snippet_start:snippet_end].strip()
                    # Clean up the display value (remove newlines, limit length)
                    secret_display = ' '.join(secret_display.split())[:100]
                    if len(secret_display) < len(value):
                        secret_display = value
                else:
                    secret_display = value

                # Validate the finding
                is_valid = True
                validation_reason = ""

                # Check for password exclusion patterns
                if any(excl in value.lower() for excl in ["password", "passwd", "pwd"]):
                    is_valid = False
                    validation_reason = "Excluded password pattern"

                # Check if it's in a comment
                if is_in_comment(context, match.start()):
                    is_valid = False
                    validation_reason = "In comment"

                # Validate credit card numbers with Luhn algorithm and test pattern detection
                if is_valid and sig["name"] == "Credit Card Number":
                    # First check if it's a test/dummy card number
                    if is_test_credit_card(value):
                        is_valid = False
                        validation_reason = "Test/dummy credit card number (repeating pattern or known test card)"
                        logging.debug(f"Credit card {value[:6]}...{value[-4:]} identified as test card")
                    # Then validate with Luhn algorithm
                    elif not luhn_checksum(value):
                        is_valid = False
                        validation_reason = "Failed Luhn checksum validation"
                        logging.debug(f"Credit card {value[:6]}...{value[-4:]} failed Luhn check")
                    # Finally check for false positive patterns using context
                    else:
                        is_false_positive, fp_reason = is_likely_not_credit_card(value, context)
                        if is_false_positive:
                            is_valid = False
                            validation_reason = f"Likely false positive: {fp_reason}"
                            logging.debug(f"Credit card {value[:6]}...{value[-4:]} rejected: {fp_reason}")

                # Validate SSN numbers using known rules
                if is_valid and sig["name"] in ["Social Security Number", "SSN Format (Context Required)"]:
                    ssn_valid, ssn_reason = validate_ssn(value)
                    if not ssn_valid:
                        is_valid = False
                        validation_reason = f"Invalid SSN: {ssn_reason}"
                        logging.debug(f"SSN failed validation: {ssn_reason}")

                # Validate AWS Access Key IDs
                if is_valid and sig["name"] == "AWS Access Key ID":
                    aws_valid, aws_reason = validate_aws_access_key(value)
                    if not aws_valid:
                        is_valid = False
                        validation_reason = f"Invalid AWS Key: {aws_reason}"
                        logging.debug(f"AWS key failed validation: {aws_reason}")

                # Validate JWT tokens
                if is_valid and sig["name"] == "JWT Token":
                    jwt_valid, jwt_reason = validate_jwt_token(value)
                    if not jwt_valid:
                        is_valid = False
                        validation_reason = f"Invalid JWT: {jwt_reason}"
                        logging.debug(f"JWT failed validation: {jwt_reason}")

                # Validate Private Keys
                if is_valid and sig["name"] == "Private Key":
                    key_valid, key_reason = validate_private_key(value)
                    if not key_valid:
                        is_valid = False
                        validation_reason = f"Invalid Private Key: {key_reason}"
                        logging.debug(f"Private key failed validation: {key_reason}")

                # Validate Database Connection Strings
                if is_valid and sig["name"] == "Database Connection String":
                    db_valid, db_reason = validate_database_url(value)
                    if not db_valid:
                        is_valid = False
                        validation_reason = f"Test DB URL: {db_reason}"
                        logging.debug(f"Database URL failed validation: {db_reason}")

                # Validate generic API tokens for entropy (catches placeholders)
                if is_valid and sig["name"] in ["API Key", "Generic Auth Token", "Slack Token", 
                                                  "SendGrid API Key", "Twilio API Key", "NPM Token"]:
                    entropy_valid, entropy_reason = validate_api_token_entropy(value)
                    if not entropy_valid:
                        is_valid = False
                        validation_reason = f"Low quality token: {entropy_reason}"
                        logging.debug(f"API token failed entropy check: {entropy_reason}")

                finding = {
                    "file_path": file_path_str,
                    "secret_type": sig["name"],
                    "secret_value": secret_display,
                    "context": context,
                    "severity": sig["severity"],
                    "line_number": text[:match.start()].count('\n') + 1,
                    "is_valid": is_valid,
                    "validation_reason": validation_reason,
                    "confidence_score": 0.9 if is_valid else 0.5
                }
                findings.append(finding)

                if is_valid:
                    logging.info(f"Valid finding: {sig['name']} in {file_path_str}")
                else:
                    logging.debug(f"Invalid finding: {sig['name']} in {file_path_str} - {validation_reason}")

        except Exception as e:
            logging.warning(f"Error scanning with signature {sig['name']}: {e}")

    return findings

def process_file(file_path, project_id, scan_session_id):
    """Process a single file for secrets with incremental scanning support.
    
    Files that haven't changed since last scan are skipped for efficiency.
    """
    file_path_str = str(file_path)
    
    logging.info(f"Starting processing of file: {file_path_str} (size: {os.path.getsize(file_path)} bytes)")

    try:
        # Check file size limit
        file_size = os.path.getsize(file_path)
        if file_size > config.scanner.max_file_size_mb * 1024 * 1024:
            logging.warning(f"File too large: {file_path_str}")
            return []

        # Calculate file hash early for change detection
        file_hash = xxhash.xxh64(open(file_path, 'rb').read()).hexdigest()
        
        # Check if file has changed since last scan (incremental scanning)
        if not file_cache_manager.is_file_changed(file_path_str, file_hash):
            logging.debug(f"File unchanged, skipping: {file_path_str}")
            return []  # File hasn't changed, skip processing
        
        logging.info(f"File changed or new, scanning: {file_path_str}")

        # Extract text from file
        text = extract_text_from_file(file_path)
        if not text:
            # Still cache the file even if no text extracted (e.g., binary)
            file_cache_manager.cache_file(file_path_str, file_hash, scan_session_id)
            return []

        # Get file metadata for change detection
        try:
            file_stat = file_path.stat()
            file_modified_at = datetime.fromtimestamp(file_stat.st_mtime)
        except Exception:
            file_modified_at = None

        # Scan text for secrets
        findings = scan_text_with_signatures(text, file_path_str)

        # Store findings in database
        stored_findings = []
        for finding in findings:
            finding_id = findings_manager.insert_finding(
                scan_session_id=scan_session_id,
                project_id=project_id,
                file_path=finding['file_path'],
                secret_type=finding['secret_type'],
                secret_value=finding['secret_value'],
                context=finding['context'],
                severity=finding['severity'],
                line_number=finding['line_number'],
                confidence_score=finding['confidence_score'],
                tool_source='custom',
                metadata={
                    'is_valid': finding['is_valid'],
                    'validation_reason': finding['validation_reason'],
                    'file_modified_at': str(file_modified_at) if file_modified_at else None,
                    'file_size': file_size
                }
            )
            if finding_id:
                stored_findings.append(finding_id)

        # Cache file processing (update cache with new hash)
        file_cache_manager.cache_file(file_path_str, file_hash, scan_session_id)

        return stored_findings

    except Exception as e:
        logging.error(f"Error processing file {file_path_str}: {e}")
        return []

def scan_directory(directory_path, project_name="default", max_workers=None):
    """Scan a directory for secrets using PostgreSQL and Redis

    Returns:
        int: Number of findings found, or -1 on error
    """

    # Initialize database connection
    if not init_database():
        logging.error("Failed to initialize database")
        return False

    # Create or get project
    project = project_manager.get_project_by_name(project_name)
    if not project:
        project_id = project_manager.create_project(
            name=project_name,
            local_path=str(directory_path),
            description=f"Scan of {directory_path}"
        )
    else:
        project_id = project['id']

    # Create scan session
    scan_session_id = scan_session_manager.create_session(
        project_id=project_id,
        scan_type='custom',
        scan_parameters={
            'directory': str(directory_path),
            'max_workers': max_workers or config.scanner.threads
        }
    )

    if not scan_session_id:
        logging.error("Failed to create scan session")
        return False

    # Update project last scan time
    project_manager.update_project_scan_time(project_id)

    # Pre-initialize OCR reader if OCR is enabled to avoid repeated initialization
    if config.scanner.enable_ocr:
        ocr_reader = get_ocr_reader()
        if ocr_reader:
            logging.info("OCR reader pre-initialized successfully")
        else:
            logging.warning("OCR initialization failed - OCR will be disabled for this scan")

    # Quick file count pass to show total progress (fast - just counting, not reading files)
    logging.info("Counting total files to scan...")
    total_files = 0
    try:
        for root, dirs, files in os.walk(directory_path):
            dirs[:] = [d for d in dirs if d not in config.scanner.excluded_paths]
            for file in files:
                file_extension = Path(file).suffix.lower()
                if (file_extension in config.scanner.supported_extensions and 
                    file_extension not in config.scanner.excluded_extensions):
                    total_files += 1
        logging.info(f"Found {total_files:,} files to scan")
    except Exception as e:
        logging.warning(f"Could not count files: {e}. Progress will show without total.")
        total_files = 0

    # Stream files and scan them directly without building a large list
    total_findings = 0
    files_processed = 0
    
    # Force single worker to avoid memory issues
    max_workers = 1
    
    # Stream scan files in very small batches to minimize memory usage
    batch_size = 1  # Process one file at a time to avoid memory issues
    current_batch = []
    
    logging.info(f"Streaming files in batches of {batch_size} with {max_workers} workers")
    logging.info("Starting file discovery and scanning...")
    
    # Use a simple counter instead of tqdm for progress to save memory
    batch_count = 0
    for root, dirs, files in os.walk(directory_path):
        # Skip excluded directories
        dirs[:] = [d for d in dirs if d not in config.scanner.excluded_paths]

        for file in files:
            file_path = Path(root) / file
            file_extension = file_path.suffix.lower()

            # Check if file type is supported and not excluded
            if (file_extension in config.scanner.supported_extensions and 
                file_extension not in config.scanner.excluded_extensions):
                
                current_batch.append(file_path)
                
                # Process batch when it's full
                if len(current_batch) >= batch_size:
                    batch_count += 1
                    logging.info(f"Processing batch {batch_count} ({len(current_batch)} files)")
                    # Log the files in this batch for debugging
                    for i, file_path in enumerate(current_batch):
                        logging.info(f"  Batch {batch_count} File {i+1}: {file_path} ({file_path.stat().st_size} bytes)")
                    
                    batch_findings = process_file_batch(current_batch, max_workers, None, project_id, scan_session_id)
                    total_findings += batch_findings
                    files_processed += len(current_batch)
                    current_batch = []
                    
                    # Update progress in Redis for real-time status display
                    try:
                        cache_manager.set(
                            'scan_progress',
                            'custom',
                            {
                                'files_processed': files_processed,
                                'total_files': total_files,
                                'total_findings': total_findings,
                                'batch_count': batch_count,
                                'status': 'running',
                                'last_update': time.time()
                            },
                            ttl_seconds=3600  # 1 hour expiry
                        )
                    except Exception:
                        pass  # Don't fail scan if Redis update fails
                    
                    # Force garbage collection after each batch
                    gc.collect()
                    logging.info(f"Batch {batch_count} complete. Total processed: {files_processed}, Total findings: {total_findings}")
    
    # Process remaining files in the last batch
    if current_batch:
        batch_count += 1
        logging.info(f"Processing final batch {batch_count} ({len(current_batch)} files)")
        batch_findings = process_file_batch(current_batch, max_workers, None, project_id, scan_session_id)
        total_findings += batch_findings
        files_processed += len(current_batch)
        gc.collect()

    logging.info(f"Scan complete. Processed {files_processed} files. Total findings: {total_findings}")
    return total_findings


def process_file_batch(batch_files, max_workers, pbar, project_id, scan_session_id):
    """Process a batch of files with ThreadPoolExecutor"""
    batch_findings = 0
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit batch tasks
        logging.info(f"Submitting {len(batch_files)} files to ThreadPoolExecutor with {max_workers} workers")
        future_to_file = {executor.submit(process_file, file_path, project_id, scan_session_id): file_path for file_path in batch_files}
        
        # Process completed tasks
        completed_count = 0
        for future in as_completed(future_to_file):
            file_path = future_to_file[future]
            completed_count += 1
            logging.info(f"Processing result {completed_count}/{len(batch_files)}: {file_path}")
            try:
                findings = future.result()
                batch_findings += len(findings) if findings else 0
                logging.info(f"File {file_path} completed successfully with {len(findings) if findings else 0} findings")
            except Exception as exc:
                logging.error(f'File {file_path} generated an exception: {exc}')
            finally:
                if pbar:
                    pbar.update(1)
    
    return batch_findings

    # Use batch processing to avoid memory overload
    max_workers = max_workers or config.scanner.threads
    batch_size = min(1000, max_workers * 50)  # Process in batches of 1000 or 50x workers
    total_findings = 0
    
    logging.info(f"Processing files in batches of {batch_size} with {max_workers} workers")

    with tqdm(total=len(files_to_scan), desc="Scanning files") as pbar:
        # Process files in batches
        for i in range(0, len(files_to_scan), batch_size):
            batch = files_to_scan[i:i + batch_size]
            batch_findings = 0
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit current batch
                future_to_file = {
                    executor.submit(process_file, file_path, project_id, scan_session_id): file_path
                    for file_path in batch
                }

                # Process batch results as they complete
                for future in as_completed(future_to_file):
                    file_path = future_to_file[future]
                    try:
                        findings = future.result()
                        batch_findings += len(findings)
                    except Exception as e:
                        logging.error(f"Error processing {file_path}: {e}")
                    pbar.update(1)
            
            total_findings += batch_findings
            # Force garbage collection after each batch
            gc.collect()
            
            logging.info(f"Completed batch {i//batch_size + 1}/{(len(files_to_scan)-1)//batch_size + 1}: {batch_findings} findings")

    # Update scan session with final results
    scan_session_manager.update_session_status(
        session_id=scan_session_id,
        status='completed',
        total_files=len(files_to_scan),
        total_findings=total_findings
    )

    logging.info(f"Scan completed: {total_findings} findings in {len(files_to_scan)} files")
    return total_findings

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="SecretSnipe Scanner")
    parser.add_argument("directory", help="Directory to scan")
    parser.add_argument("--project", default="default", help="Project name")
    parser.add_argument("--workers", type=int, help="Number of worker threads")
    parser.add_argument("--config", help="Configuration file path")

    args = parser.parse_args()

    # Load configuration
    if args.config:
        # Load custom config if specified
        pass

    # Load signatures
    if not load_signatures():
        logging.error("Failed to load signatures")
        return 1

    # Initialize Redis
    if not redis_manager or not redis_manager.ping():
        logging.warning("Redis not available - continuing without caching")

    # Scan directory
    success = scan_directory(
        directory_path=Path(args.directory),
        project_name=args.project,
        max_workers=args.workers
    )

    return 0 if success >= 0 else 1

if __name__ == "__main__":
    sys.exit(main())